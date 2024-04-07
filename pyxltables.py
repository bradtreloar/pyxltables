
import re
from typing import Any, Dict, List
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


TableRow = Dict[str, Any]


def load_rows(filepath, worksheet_name: str | None = None):
    """
    Read data from XLSX table.
    """
    workbook: Workbook = load_workbook(filename=filepath, read_only=True)

    # Use the first worksheet by default.
    if worksheet_name is None:
        worksheet_name = str(workbook.get_sheet_names()[0])
    worksheet: Worksheet = workbook.get_sheet_by_name(worksheet_name)
    rows = read_rows(worksheet)
    workbook.close()
    return rows


def read_rows(worksheet: Worksheet):
    """
    Read rows from worksheet.
    """
    rows: List[TableRow] = []
    fieldnames = get_fieldnames(worksheet)
    for row in worksheet.iter_rows(min_row=2):
        if row[0].value is None:
            return rows
        values: TableRow = {}
        for i, fieldname in enumerate(fieldnames):
            values[fieldname] = row[i].value
        rows.append(values)
    return rows


def get_fieldnames(worksheet: Worksheet):
    """
    Parse snakecase fieldnames from first row of sheet.
    """
    fieldnames: List[str] = []
    col_index = 0
    reached_last_column = False
    while not reached_last_column:
        value = worksheet.cell(1, col_index + 1).value
        if value is None:
            reached_last_column = True
        else:
            fieldnames.append(snakecase(value))
            col_index += 1
    return fieldnames


def snakecase(value: str):
    """
    Converts table header to snakecase.
    """
    value = value.strip()
    value = value.lower()
    value = re.sub(r"[/ =:-]", "_", value)
    value = re.sub(r"[\(\)\.]", "", value)
    value = value.replace("___", "_")
    value = value.replace("__", "_")
    value = value.strip("_")
    return value
